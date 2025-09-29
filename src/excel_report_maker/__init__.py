# excel_report_generator.py

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font

class ExcelReportGenerator:
    def __init__(self, results, intro_text):
        """
        Initializes the Excel report generator.

        Args:
            results (dict): A dictionary where each key is a sheet name (derived from a SQL file name)
                            and each value is a dictionary mapping query titles to DataFrames.
            intro_text (list of str): A list of text lines to be added to the introduction sheet.
        """
        self.results = results
        self.intro_text = intro_text
        self.wb = Workbook()
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        self.global_table_counter = 1

    def create_introduction_sheet(self):
        """
        Creates the Introduction worksheet and populates it with the intro text.
        """
        ws = self.wb.create_sheet("Introduction", 0)
        ws.append(["Introduction"])
        ws["A1"].font = Font(bold=True, size=14)
        
        for line in self.intro_text:
            ws.append([line])
        
        ws.column_dimensions['A'].width = 100

    def append_df_as_table(self, ws, df, table_title, start_row):
        """
        Appends a Pandas DataFrame to a worksheet as a formatted table.

        Args:
            ws (Worksheet): The worksheet object.
            df (DataFrame): The Pandas DataFrame to add.
            table_title (str): Title to display above the table.
            start_row (int): The starting row number in the worksheet.

        Returns:
            int: The row number after the inserted table.
        """
        ws.cell(row=start_row, column=1, value=table_title).font = Font(bold=True, size=12)
        start_row += 1
        initial_data_row = start_row

        for i, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=0):
            ws.append(row)
        rows_appended = i + 1
        end_row = start_row + rows_appended - 1
        start_col = 1
        end_col = df.shape[1]

        table_ref = f"{get_column_letter(start_col)}{initial_data_row}:{get_column_letter(end_col)}{end_row}"
        table = Table(displayName=f"Table{self.global_table_counter}", ref=table_ref)
        self.global_table_counter += 1

        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        rate_cols = [col for col in df.columns if "Rate" in col]
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name in rate_cols:
                for row_idx in range(initial_data_row + 1, end_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = '0%'

        return end_row + 2

    def add_image_from_figure(self, ws, fig, start_row, start_col='A', width=600, height=400):
        """
        Add matplotlib figure as image to worksheet.
        
        Args:
            ws (Worksheet): The worksheet object
            fig: Matplotlib figure object
            start_row (int): Row to place image
            start_col (str or int): Column to place image (default 'A')
            width (int): Image width in pixels (default 600)
            height (int): Image height in pixels (default 400)
            
        Returns:
            int: Row number after the image
        """
        try:
            from openpyxl.drawing.image import Image
            import io
            
            img_buffer = io.BytesIO()
            fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            img_buffer.seek(0)
            
            img = Image(img_buffer)
            img.width = width
            img.height = height
            
            if isinstance(start_col, str):
                anchor_cell = f"{start_col}{start_row}"
            else:
                anchor_cell = f"{get_column_letter(start_col)}{start_row}"
            
            img.anchor = anchor_cell
            ws.add_image(img)
            
            rows_occupied = int(height / 15) + 2
            return start_row + rows_occupied
            
        except ImportError:
            print("Warning: Image embedding requires matplotlib. Skipping image.")
            return start_row + 1

    def add_image_from_file(self, ws, image_path, start_row, start_col='A', width=600, height=400):
        """
        Add PNG file as image to worksheet.
        
        Args:
            ws (Worksheet): The worksheet object
            image_path (str): Path to PNG file
            start_row (int): Row to place image
            start_col (str or int): Column to place image
            width (int): Image width in pixels
            height (int): Image height in pixels
            
        Returns:
            int: Row number after the image
        """
        try:
            from openpyxl.drawing.image import Image
            from pathlib import Path
            
            if not Path(image_path).exists():
                print(f"Warning: Image file not found: {image_path}")
                return start_row + 1
            
            img = Image(str(image_path))
            img.width = width
            img.height = height
            
            if isinstance(start_col, str):
                anchor_cell = f"{start_col}{start_row}"
            else:
                anchor_cell = f"{get_column_letter(start_col)}{start_row}"
            
            img.anchor = anchor_cell
            ws.add_image(img)
            
            rows_occupied = int(height / 15) + 2
            return start_row + rows_occupied
            
        except ImportError:
            print("Warning: Image embedding not available. Skipping image.")
            return start_row + 1

    def create_results_sheets(self):
        """
        Creates a separate worksheet for each SQL file (based on the file name) and
        populates it with all tables derived from that file's queries.
        Assumes that self.results is a dictionary with sheet names as keys and values
        that are dictionaries mapping query titles to DataFrames.
        """
        for sheet_name, queries_dict in self.results.items():
            ws = self.wb.create_sheet(sheet_name)
            current_row = 1

            for table_title, df in queries_dict.items():
                current_row = self.append_df_as_table(ws, df, table_title, current_row)

            for col in ws.columns:
                max_length = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
                adjusted_width = max_length + 2
                ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    def create_results_sheets_with_images(self, image_functions=None):
        """
        Creates results sheets with optional image embedding after each table.
        
        Args:
            image_functions (dict): Optional dictionary mapping sheet names to image generation functions.
                                   Functions should accept (df, table_title) and return matplotlib figure(s)
                                   or list of figures.
        """
        for sheet_name, queries_dict in self.results.items():
            ws = self.wb.create_sheet(sheet_name)
            current_row = 1

            for table_title, df in queries_dict.items():
                current_row = self.append_df_as_table(ws, df, table_title, current_row)
                
                if image_functions and sheet_name in image_functions:
                    try:
                        image_func = image_functions[sheet_name]
                        figures = image_func(df, table_title)
                        
                        if not isinstance(figures, list):
                            figures = [figures]
                        
                        for fig in figures:
                            if fig is not None:
                                current_row = self.add_image_from_figure(ws, fig, current_row)
                                current_row += 1
                                
                                try:
                                    import matplotlib.pyplot as plt
                                    plt.close(fig)
                                except:
                                    pass
                    except Exception as e:
                        print(f"Warning: Failed to add images for {sheet_name}/{table_title}: {e}")

            for col in ws.columns:
                max_length = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
                adjusted_width = min(max_length + 2, 50)  # Cap width at 50
                ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    def generate_workbook(self, output_path, image_functions=None):
        """
        Creates the workbook with the Introduction and all results sheets and saves it.

        Args:
            output_path (str): The file path where the workbook will be saved.
            image_functions (dict): Optional dictionary of image generation functions for embedding charts.
        """
        self.create_introduction_sheet()
        
        if image_functions:
            self.create_results_sheets_with_images(image_functions)
        else:
            self.create_results_sheets()
            
        self.wb.save(output_path)
        print(f"Workbook successfully saved to {output_path}")

if __name__ == '__main__':
    print('Welcome to excel-report-maker!')
    print('Import ExcelReportGenerator into your packages to start quickly generating report files!')
    print('New in this version: Image embedding support with add_image_from_figure() and add_image_from_file()')